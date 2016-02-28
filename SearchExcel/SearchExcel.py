from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import openpyxl
import os.path

def CheckIfBookExists(bookName):
        path = './' + bookName + '.xlsx'
        if(os.path.isfile(path)):
                return True
        else:
                print("File does not exist")
                return False
        
def Main():
        print("Dan's Excel Workbook Data Sorter V1")
        workBookName = input("Enter workbook name: ")
        while(CheckIfBookExists(workBookName) == False):
                workBookName = input("Enter workbook name: ")
        workBookName += '.xlsx'
        print(workBookName)
        wb = load_workbook(filename = workBookName)
        
        #sheetName = input("Please enter the sheet name: ")
        #while(CheckIfSheetExists(wb, sheetName) == False):
              #  workBookName = input("Please enter the sheet name: ")


        #sheet = wb.get_sheet_by_name('TestData')
        sheet = wb.active #gets the active sheet

        coloumn = input("Please input the coloumn letter you wish to parse.")

        rowData = {}
        for row in range(1, sheet.max_row + 1):
                dataEntered = sheet[coloumn + str(row)].value
                rowData['L' + str(row)] = dataEntered

        keyword = input("PLease enter the keyword you wish to sort by: ")          
        Search(rowData, keyword)


def Search(data, keyword):
        print("Seaching...")
        searchedData = {}
        for row in range(1, len(data) + 1):
                line = data['L' + str(row)]
                #print(line)
                if(line.find(keyword == -1)):    #returns a -1 if not found
                        print("No match.")
                        searchedData[len(searchedData)+1] = 'NOT RELAVENT TO SEARCH!'
                else:
                    searchedData[len(searchedData)+1] = line
                    print("Found Match.")
     
        WriteToNewBook(searchedData, keyword)


def WriteToNewBook(data, keyword):
        newWb = openpyxl.Workbook()
        sheet = newWb.get_sheet_by_name('Sheet')
        sheet.title = keyword
        #print(sheet.title)
        for row in range(1, len(data) + 1):
                #print(type(row))
                sheet['A' + str(row)] = data[row]

        title = keyword + 'ParsedData.xlsx'
        newWb.save(title)

        
Main()
